import csv

import openpyxl
from django.conf import settings
from django.core.files import temp as tempfile
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


def csv_response(filename, table):
    """Return a CSV file of the given table as an HttpResponse.

    Args:

        filename: the name of the downloaded CSV file. The extension will be
            '.csv'. This parameter is inserted directly to the response's
            Content-Disposition, and must be escaped accordingly.

        table: a 2-dimensional iterable, in row-major order.

    Returns:

        A CSV HttpResponse with appropriate content_type and
        Content-Disposition.

    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="%s.csv"' % filename
    writer = csv.writer(response)
    for row in table:
        # Convert generators to lists for use by writer.writerow.
        writer.writerow(list(row))
    return response


def xlsx_response(filename, table, max_width=118, max_height=90):
    """Return a Microsoft Excel 2007+ file of the given table as an
     HttpResponse.

    Args:

        filename: the name of the downloaded file. The extension will be
        '.xlsx'. This parameter is inserted directly to the response's
        Content-Disposition, and must be escaped accordingly.

        table: a 2-dimensional iterable, in row-major order.

    Returns:

        A Microsoft Excel 2007+ HttpResponse with appropriate content_type and
        Content-Disposition.

    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % filename
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    cell_widths = dict()
    cell_heights = dict()

    for r, row in enumerate(table, start=1):
        for c, cell in enumerate(row, start=1):
            ws_cell = worksheet.cell(row=r, column=c)
            ws_cell.value = cell

            if type(cell) in [str, unicode]:
                cell_str = ws_cell.value.encode('utf-8')
            elif type(cell) in [float]:
                ws_cell.number_format = '0.00'
                cell_str = str(ws_cell.value)
            else:
                cell_str = str(cell)

            cell_widths[c] = min(max((cell_widths.get(c, 0), len(cell_str))), max_width)
            cell_height = int(len(cell_str.split('\n')) * 15)
            cell_heights[r] = min(max((cell_heights.get(r, 0), cell_height)), max_height)

    for column, width in cell_widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width + 1

    for row, height in cell_heights.items():
        worksheet.row_dimensions[row].height = height

    # Save to temporary file
    if settings.FILE_UPLOAD_TEMP_DIR:
        my_temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', dir=settings.FILE_UPLOAD_TEMP_DIR)
    else:
        my_temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx')

    workbook.save(my_temp_file.name)
    my_file = my_temp_file.file
    response.write(my_file.read())
    my_file.close()
    return response


# TODO: Add ability to handle multiple worksheets
